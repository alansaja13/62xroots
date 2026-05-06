"""
Importador GrowLog.xlsx → Django DB
Correr con: $env:PYTHONIOENCODING="utf-8"; py -3 import_excel.py
"""
import os, sys
from datetime import datetime, date, timedelta, time as dtime
from decimal import Decimal

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
import django
django.setup()

import openpyxl
from django.utils import timezone
from growlog.models import (
    Cultivo, Planta, MedicionAmbiente, Nutriente, Riego,
    NutrienteAplicado, Evento, MedicionPlanta, Tarea,
)

XLSX = r'C:\Users\Alan\Documents\Alan\Cultivan2\GrowLog.xlsx'
wb = openpyxl.load_workbook(XLSX, data_only=True)

def log(msg): print(msg)

# ── helpers ──────────────────────────────────────────────────────────────────

def to_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    if isinstance(val, int): return date(1899, 12, 30) + timedelta(days=val)
    return None

def to_time(val):
    if isinstance(val, dtime): return dtime(val.hour, val.minute, val.second)
    if isinstance(val, str) and ':' in val:
        p = val.split(':')
        return dtime(int(p[0]), int(p[1]), int(float(p[2])) if len(p) > 2 else 0)
    return dtime(12, 0, 0)

def make_dt(fecha_val, hora_val=None):
    d = to_date(fecha_val)
    t = to_time(hora_val) if hora_val is not None else dtime(12, 0, 0)
    return timezone.make_aware(datetime.combine(d, t))

def dec(val):
    if val is None: return None
    return Decimal(str(round(float(val), 4)))

# ── Cultivo ───────────────────────────────────────────────────────────────────

cultivo = Cultivo.objects.filter(nombre="Fancy Gummy + Satélite #1").first()
if not cultivo:
    sys.exit("ERROR: Cultivo no encontrado. Corré seed.py primero.")

cultivo.fecha_inicio = date(2026, 4, 23)
cultivo.lampara_modelo = "COB LED 300W"
cultivo.lampara_watts_reales = 300
cultivo.sustrato = "BioBizz Light-Mix + 30% perlita"
cultivo.carpa_dimensiones = "80×80×200 cm"
cultivo.notas = "6 plantas: 5x Fancy Gummy + 1x Satélite. SCRoG planificado post-topping."
cultivo.save()
log(f"✓ Cultivo actualizado: {cultivo.nombre} — inicio {cultivo.fecha_inicio}")

# ── Plantas ───────────────────────────────────────────────────────────────────

planta_map = {}  # apodo → Planta

# Archivar FG-3 (no existe en el cultivo real — el seed la creó por error)
fg3 = Planta.objects.filter(cultivo=cultivo, apodo__in=["FG3", "FG-3"]).first()
if fg3:
    fg3.archivado = True
    fg3.estado = "descartada"
    fg3.save()
    log(f"✓ {fg3.apodo} archivada (planta inexistente en el cultivo real)")

# Renombrar SAT-1 → Satelite
for apodo_viejo in ["SAT-1", "Satelite"]:
    sat = Planta.objects.filter(cultivo=cultivo, apodo=apodo_viejo).first()
    if sat:
        sat.apodo = "Satelite"
        sat.strain = "Satélite"
        sat.posicion_tent = "abajo_izq"
        sat.notas_genetica = (
            "Sem 1: 6–7 nudos. Vino avanzada de ONG, ya en veg activo al trasplante. "
            "NO se toppea — ya tiene 2 colas naturales. Solo LST para mantener en SCRoG. "
            "Tallo con coloración rojiza (probable genética). Hoja basal amarilla (senescencia normal). "
            "Posición esquina abajo-izquierda elegida para cosecha temprana (45d flo)."
        )
        sat.save()
        log(f"✓ {apodo_viejo} → Satelite (abajo izq)")
        planta_map["Satelite"] = sat
        break

# Corregir posiciones, renombrar (quitar guión) y agregar notas de las FG
# El seed las creó como "FG-1", el Excel las nombra "FG1"
plantas_config = {
    "FG-1": ("FG1", "centro",     "Sem 1: 5–6 nudos. Lista para topping día 6–7. Estructura limpia, sin síntomas. Toppeada día 10 (03/05)."),
    "FG-2": ("FG2", "arriba_izq", "Sem 1: 4–5 nudos. Topping día 10–12. Desarrollo normal. Toppeada día 10 (03/05)."),
    "FG-4": ("FG4", "abajo_der",  "Sem 1: 5–6 nudos. Lista para topping día 6–7. Estructura limpia. LST día 8 (01/05): apical derecha guiada OK; apical izquierda fracturada accidentalmente — entablillada con scotch + alambre. Revisar 08/05. Manchas: splash damage (no patógeno)."),
    "FG-5": ("FG5", "arriba_der", "Sem 1: 4–5 nudos. Topping ~día 10–12. Tallo algo más fino que las demás. Hoja basal amarilla (senescencia normal). Toppeada día 10 (03/05)."),
    # Por si ya se renombraron en una corrida anterior:
    "FG1":  ("FG1", "centro",     None),
    "FG2":  ("FG2", "arriba_izq", None),
    "FG4":  ("FG4", "abajo_der",  None),
    "FG5":  ("FG5", "arriba_der", None),
}

for apodo_viejo, (apodo_nuevo, posicion, notas) in plantas_config.items():
    p = Planta.objects.filter(cultivo=cultivo, apodo=apodo_viejo).first()
    if p:
        p.apodo = apodo_nuevo
        p.posicion_tent = posicion
        if notas:
            p.notas_genetica = notas
        p.save()
        planta_map[apodo_nuevo] = p
        if apodo_viejo != apodo_nuevo:
            log(f"✓ {apodo_viejo} → {apodo_nuevo} ({posicion})")
        else:
            log(f"✓ {apodo_nuevo} posición actualizada ({posicion})")

log(f"✓ Mapa de plantas: {list(planta_map.keys())}")

# ── Nutrientes Greenhouse ──────────────────────────────────────────────────────

nut_gh = [
    dict(nombre="Powder Feeding GROW",    marca="Greenhouse", npk="25-2,8-6,9",
         etapa_recomendada="veg",    dosis_recomendada_min=dec("0.50"), dosis_recomendada_max=dec("0.66")),
    dict(nombre="Powder Feeding HYBRIDS", marca="Greenhouse", npk="15-7-22",
         etapa_recomendada="flora",  dosis_recomendada_min=dec("0.66"), dosis_recomendada_max=dec("0.66")),
    dict(nombre="Additive Feeding BOOSTER", marca="Greenhouse", npk="0-30-27",
         etapa_recomendada="aditivo", dosis_recomendada_min=dec("0.33"), dosis_recomendada_max=dec("0.33")),
]
nut_map = {}
for nd in nut_gh:
    n, created = Nutriente.objects.get_or_create(
        nombre=nd["nombre"], marca=nd["marca"], defaults=nd
    )
    nut_map[nd["nombre"]] = n
    log(f"{'✓ Creado' if created else '→ Existe'}: {n}")

def nutriente_de_str(nut_str, dosis_val):
    """Retorna (Nutriente, Decimal dosis) o (None, None)."""
    if not nut_str or "sin" in str(nut_str).lower():
        return None, None
    s = str(nut_str)
    if "GROW" in s:
        return nut_map["Powder Feeding GROW"], dec(dosis_val) or dec("0.5")
    if "HYBRIDS" in s:
        return nut_map["Powder Feeding HYBRIDS"], dec(dosis_val) or dec("0.66")
    if "BOOSTER" in s:
        return nut_map["Additive Feeding BOOSTER"], dec(dosis_val) or dec("0.33")
    return None, None

# ── Limpiar datos demo del seed ────────────────────────────────────────────────

MedicionAmbiente.objects.filter(cultivo=cultivo, notas="Primera medición — seed inicial").delete()
Tarea.objects.filter(cultivo=cultivo, titulo__in=[
    "Revisar pH del agua próximo riego",
    "Medir altura de las plantas",
    "Verificar distancia de lámpara",
]).delete()
log("✓ Datos demo del seed eliminados")

# ── Registro Diario → MedicionAmbiente + Riego ────────────────────────────────

ws_rd = wb['Registro Diario']
med_count = 0
riego_count = 0

for row in ws_rd.iter_rows(min_row=5, values_only=True):
    if not any(c is not None for c in row):
        continue

    fecha_val  = row[0]
    hora_val   = row[1]
    # row[2] = Día Cultivo, row[3] = Etapa
    temp       = row[4]
    hr         = row[5]
    # row[6] = VPD (calculado, ignorar)
    rego       = row[7]
    vol_planta = row[8]
    ph         = row[9]
    nut_str    = row[10]
    ec         = row[11]
    dosis      = row[12]
    aspecto    = row[13]
    sintomas   = row[14]
    altura     = row[15]
    accion     = row[16]
    prox_acc   = row[17]
    obs        = row[18] if len(row) > 18 else None

    if fecha_val is None:
        continue

    # Notas combinadas
    partes = []
    if obs:      partes.append(str(obs))
    if accion:   partes.append(f"Acción: {accion}")
    if prox_acc: partes.append(f"Próx: {prox_acc}")
    if sintomas and str(sintomas).strip().lower() not in ("ninguno", "ninguno particular", ""):
        partes.append(f"Síntomas: {sintomas}")
    notas = " | ".join(partes)

    ts = make_dt(fecha_val, hora_val)

    # MedicionAmbiente — solo si hay temp y HR
    if temp is not None and hr is not None:
        _, created = MedicionAmbiente.objects.get_or_create(
            cultivo=cultivo,
            timestamp=ts,
            defaults={
                "temperatura_c": dec(temp),
                "humedad_relativa": dec(hr),
                "notas": notas,
            }
        )
        if created:
            med_count += 1

    # Riego
    if str(rego or "").strip().lower() in ("si", "sí") and vol_planta:
        vol_total = int(float(vol_planta) * 5)
        riego, r_created = Riego.objects.get_or_create(
            cultivo=cultivo,
            timestamp=ts,
            defaults={
                "volumen_total_ml": vol_total,
                "volumen_por_planta_ml": int(float(vol_planta)),
                "ph_agua": dec(ph),
                "ec_solucion": dec(ec),
                "notas": notas,
            }
        )
        if r_created:
            riego_count += 1
            nut, dosis_dec = nutriente_de_str(nut_str, dosis)
            if nut:
                NutrienteAplicado.objects.get_or_create(
                    riego=riego, nutriente=nut,
                    defaults={"dosis_g_por_litro": dosis_dec}
                )

    # Día sin temp/HR → ignorar (ej: día 12 sin datos)

log(f"✓ Mediciones de ambiente creadas: {med_count}")
log(f"✓ Riegos creados: {riego_count}")

# ── Eventos ───────────────────────────────────────────────────────────────────

TIPO_MAP = {
    "inicio cultivo":         "transplante",
    "falla técnica":          "problema",
    "falla tecnica":          "problema",
    "ajuste ambiente":        "ajuste_ambiente",
    "ajuste lámpara":         "ajuste_ambiente",
    "ajuste lampara":         "ajuste_ambiente",
    "primer riego":           "otro",
    "cambio posicion de plantas": "otro",
    "planificación training": "otro",
    "planificacion training": "otro",
    "riego + primer fertirriego": "otro",
    "riego + fertirriego":    "otro",
    "lst":                    "lst",
    "lst + accidente":        "lst",
    "defoliación":            "defoliacion",
    "defoliacion":            "defoliacion",
    "diagnóstico":            "diagnostico",
    "diagnostico":            "diagnostico",
    "remonte post-lst":       "otro",
    "topping":                "topping",
    "cambio ciclo":           "cambio_ciclo",
}

def parse_plantas_afectadas(val):
    if val is None or isinstance(val, (datetime, date)):
        return []
    s = str(val).strip()
    if not s or s == "—":
        return []
    if "todas" in s.lower() or "5/5" in s:
        return list(planta_map.values())
    resultado = []
    for nombre in s.replace(" y ", ",").split(","):
        n = nombre.strip()
        if n in planta_map:
            resultado.append(planta_map[n])
    return resultado

ws_ev = wb['Eventos']
ev_count = 0

for row in ws_ev.iter_rows(min_row=4, values_only=True):
    if not any(c is not None for c in row):
        continue

    fecha_val  = row[0]
    hora_val   = row[1]
    dia        = row[2]
    tipo_raw   = row[3]
    desc_raw   = row[4]
    plantas_raw= row[5]
    resultado  = row[6]
    followup   = row[7]

    if fecha_val is None:
        continue

    # Detectar desplazamiento de columnas (cuando dia tiene el tipo en vez del número)
    if isinstance(dia, str) and not str(dia).strip().isdigit():
        tipo_str   = str(dia).strip()
        plantas_str = tipo_raw
        desc       = desc_raw
    else:
        tipo_str   = str(tipo_raw or "").strip()
        plantas_str = plantas_raw
        desc       = desc_raw

    if not desc:
        continue

    tipo_key = tipo_str.lower()
    tipo_db  = TIPO_MAP.get(tipo_key, "otro")

    ts = make_dt(fecha_val, hora_val if not isinstance(hora_val, (int, float)) else None)

    # Follow-up
    fu_fecha = None
    fu_desc  = ""

    # Si plantas_raw es un datetime, es realmente la fecha de follow-up
    if isinstance(plantas_raw, datetime):
        fu_fecha = plantas_raw.date()
    if isinstance(plantas_str, datetime):
        fu_fecha = plantas_str.date()

    # Caso especial: fractura FG4 → follow-up el 08/05
    if "fractura" in str(desc).lower() or "accidente" in tipo_key:
        fu_fecha = date(2026, 5, 8)
        fu_desc  = "Revisar fractura FG4: si el tallo está rígido y sin marchitez, retirar scotch. Si colapso, evaluar poda."

    if followup and isinstance(followup, str):
        fu_desc = (fu_desc + " | " + followup).strip(" | ") if fu_desc else followup

    # Agregar resultado a la descripción
    desc_completa = str(desc)
    if resultado and isinstance(resultado, str) and not resultado.startswith("Pendiente"):
        desc_completa += f"\nResultado: {resultado}"

    ev, created = Evento.objects.get_or_create(
        cultivo=cultivo,
        timestamp=ts,
        tipo=tipo_db,
        defaults={
            "descripcion": desc_completa,
            "follow_up_fecha": fu_fecha,
            "follow_up_descripcion": fu_desc,
            "follow_up_resuelto": False,
        }
    )

    if created:
        ev_count += 1
        plantas = parse_plantas_afectadas(plantas_str)
        if plantas:
            ev.plantas_afectadas.set(plantas)

log(f"✓ Eventos creados: {ev_count}")

# ── MedicionPlanta — semana 1 ─────────────────────────────────────────────────

semana1_data = {
    "FG1":     {"altura_cm": dec("31"), "nudos_count": 6},
    "FG2":     {"altura_cm": dec("32"), "nudos_count": 5},
    "Satelite":{"altura_cm": dec("42"), "nudos_count": 7},
    "FG4":     {"altura_cm": dec("34"), "nudos_count": 6},
    "FG5":     {"altura_cm": dec("33"), "nudos_count": 5},
}
fecha_s1 = date(2026, 4, 30)
for apodo, vals in semana1_data.items():
    p = planta_map.get(apodo)
    if p:
        MedicionPlanta.objects.get_or_create(
            planta=p, fecha=fecha_s1,
            defaults={**vals, "aspecto_general": "bueno", "sintomas": ""},
        )
log("✓ Mediciones de planta semana 1 cargadas")

# ── Tareas desde Próximos Pasos ───────────────────────────────────────────────

TAREAS = [
    # (titulo, categoria, prioridad, fecha_obj, completada, ts_completada, descripcion)
    ("Revisar fractura FG4", "observacion", "urgente", date(2026, 5, 8), False, None,
     "Verificar si la entablilladura con scotch + alambre sostuvo. Si tallo rígido sin marchitez → retirar scotch. Si colapso → evaluar podar la rama."),
    ("Comprar medidor EC/TDS", "compra", "urgente", None, False, None,
     "Necesario antes del próximo fertirriego. Buscar en MercadoLibre."),
    ("Comprar ventilador clip 220V", "compra", "normal", None, False, None,
     "Mejorar circulación interna del tent. Clip para colgar en estructura."),
    ("Instalar malla SCRoG", "training", "normal", date(2026, 5, 13), False, None,
     "Cuando ramas laterales post-topping tengan 10-15cm. Altura 40-45cm del piso. Cuadros de 5-6cm."),
    ("Tucking inicial SCRoG", "training", "baja", None, False, None,
     "Pasar ramas por debajo de la malla hacia los lados, llenando todos los cuadros. Repetir cada 2-3 días."),
    ("Observar recuperación post-topping", "observacion", "normal", date(2026, 5, 6), False, None,
     "Plantas pueden verse tristes 24-48hs. Normal. NO regar fuerte, NO fertilizar, NO defoliar."),
    # Completadas
    ("LST suave a Satelite", "training", "normal", date(2026, 5, 1), True,
     timezone.make_aware(datetime(2026, 5, 1, 21, 0)),
     "LST completado: 3 colas (2 apicales + 1 lateral) guiadas con alambre forrado. Defoliación base realizada. Remonte positivo al día siguiente."),
    ("Topping FG1, FG2, FG5", "training", "urgente", date(2026, 5, 3), True,
     timezone.make_aware(datetime(2026, 5, 3, 11, 30)),
     "Topping ejecutado sobre 5° nudo. FG4 con LST: apical derecha guiada, apical izquierda fracturada y entablillada."),
    ("Bajar lámpara post-topping", "mantenimiento", "normal", date(2026, 5, 3), True,
     timezone.make_aware(datetime(2026, 5, 3, 12, 0)),
     "Compensar pérdida de meristemo dominante. Las laterales necesitan luz directa."),
]

t_count = 0
for titulo, cat, prio, fecha_obj, completada, ts_comp, desc in TAREAS:
    _, created = Tarea.objects.get_or_create(
        cultivo=cultivo, titulo=titulo,
        defaults={
            "categoria": cat, "prioridad": prio,
            "fecha_objetivo": fecha_obj,
            "completada": completada, "completada_en": ts_comp,
            "descripcion": desc,
        }
    )
    if created:
        t_count += 1

log(f"✓ Tareas creadas: {t_count}")

# ── Resumen final ─────────────────────────────────────────────────────────────

log("\n" + "="*50)
log(f"Cultivo:     {cultivo.nombre}")
log(f"Mediciones:  {MedicionAmbiente.objects.filter(cultivo=cultivo).count()}")
log(f"Riegos:      {Riego.objects.filter(cultivo=cultivo).count()}")
log(f"Eventos:     {Evento.objects.filter(cultivo=cultivo).count()}")
log(f"Tareas:      {Tarea.objects.filter(cultivo=cultivo).count()}")
log(f"Plantas:     {Planta.objects.filter(cultivo=cultivo, archivado=False).count()} activas")
log("="*50)
log("🌱 Import completo!")
